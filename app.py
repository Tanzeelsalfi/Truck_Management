import sqlite3
import json
import os
import datetime # For system timestamp
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory
from cryptography.fernet import Fernet, InvalidToken
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

app = Flask(__name__)
app.secret_key = "your_very_secret_key_for_flashing_messages_v7" # IMPORTANT: Change this!

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_FILE = os.path.join(BASE_DIR, "receipt_scans.db")
ENCRYPTION_KEY = b"NWRmMTk3ZWUwY2RjNjA3NWY4NzQ2NmQyOGRkYzczMmM="
fernet = Fernet(ENCRYPTION_KEY)

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    # Scans table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS scans (
            serial TEXT PRIMARY KEY,
            timestamp TEXT,       -- Timestamp from QR
            truck_number TEXT,
            system_timestamp TEXT -- Timestamp of when the scan was recorded by the server
        )
    """)
    # Check and add columns if they don't exist (for existing DBs)
    cur.execute("PRAGMA table_info(scans)")
    columns = [column[1] for column in cur.fetchall()]
    if 'truck_number' not in columns:
        try:
            cur.execute("ALTER TABLE scans ADD COLUMN truck_number TEXT")
            # print("INFO: Added 'truck_number' column to 'scans' table.")
        except sqlite3.OperationalError as e:
            print(f"INFO/WARNING: Could not add 'truck_number' to 'scans' (may already exist): {e}")
    if 'system_timestamp' not in columns:
        try:
            cur.execute("ALTER TABLE scans ADD COLUMN system_timestamp TEXT")
            # print("INFO: Added 'system_timestamp' column to 'scans' table.")
        except sqlite3.OperationalError as e:
            print(f"INFO/WARNING: Could not add 'system_timestamp' to 'scans' (may already exist): {e}")

    # Truck Master List table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS truck_master_list (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            truck_number TEXT UNIQUE NOT NULL
        )
    """)
    # print("INFO: Database tables checked/created.")
    conn.commit()
    conn.close()
init_db()

def get_truck_master_list_from_db():
    conn = sqlite3.connect(DB_FILE); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    cur.execute("SELECT id, truck_number FROM truck_master_list ORDER BY truck_number ASC")
    trucks = cur.fetchall(); conn.close(); return trucks

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route("/", methods=["GET"])
def index():
    truck_master_list = get_truck_master_list_from_db()
    # Get previously submitted truck info from session or query params if needed for extreme stickiness,
    # but for now, rely on Flask passing it back after POST for simplicity.
    return render_template("index.html",
                           result_message=request.args.get('result_message'), # For GET redirects with messages
                           error_message=request.args.get('error_message'),   # For GET redirects
                           truck_numbers_master=truck_master_list,
                           submitted_truck_select=request.args.get('truck_select', ''), # For GET redirects
                           submitted_truck_other=request.args.get('truck_other', '')     # For GET redirects
                           )

@app.route("/verify", methods=["POST"])
def verify():
    encrypted_data_from_form = request.form.get("encrypted", "").strip()
    selected_truck = request.form.get("truck_number_select", "").strip()
    other_truck_number = request.form.get("truck_number_other", "").strip()
    
    final_truck_number_for_scan = ""
    if selected_truck == "_other_":
        final_truck_number_for_scan = other_truck_number
    elif selected_truck:
        final_truck_number_for_scan = selected_truck
    
    result_val = None
    error_val = None

    if encrypted_data_from_form:
        try:
            decrypted_data = fernet.decrypt(encrypted_data_from_form.encode()).decode()
            if "|" not in decrypted_data: 
                raise ValueError("Format error: Decrypted data does not contain '|'.")
            
            serial, qr_timestamp = decrypted_data.split("|", 1)

            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute("SELECT timestamp, truck_number, system_timestamp FROM scans WHERE serial = ?", (serial,))
            existing_scan = cur.fetchone()

            if existing_scan:
                ex_qr_ts, ex_truck_db, ex_sys_ts = existing_scan
                truck_info = f"Truck '{ex_truck_db}'" if ex_truck_db and ex_truck_db.strip() else "an unspecified truck"
                system_ts_info = f" (recorded on {ex_sys_ts})" if ex_sys_ts else ""
                error_val = f"❌ Duplicate! Redeemed by {truck_info} on {ex_qr_ts}{system_ts_info}."
            else:
                cts = final_truck_number_for_scan if final_truck_number_for_scan else None
                current_system_timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                cur.execute("INSERT INTO scans (serial, timestamp, truck_number, system_timestamp) VALUES (?, ?, ?, ?)",
                            (serial, qr_timestamp, cts, current_system_timestamp))
                conn.commit()
                
                success_message = f"✅ Verified:\nSerial No: {serial}\nQR Timestamp: {qr_timestamp}\nSystem Timestamp: {current_system_timestamp}"
                if cts:
                    success_message += f"\nTruck No: {cts}"
                result_val = success_message
            conn.close()
        except InvalidToken: error_val = "❌ Invalid/Tampered data."
        except ValueError as ve: error_val = f"❌ Malformed decrypted data: {ve}"
        except Exception as e: error_val = f"❌ Unexpected error: {type(e).__name__} - {e}"
    else: error_val = "❌ No QR data."
    
    truck_master_list = get_truck_master_list_from_db()
    # Instead of rendering directly, redirect back to index with messages and truck info as query params
    # This helps with the page refresh behavior and keeping the URL clean for GET requests.
    # However, for modal display, passing directly on re-render is fine as we did.
    # For true stickiness across unrelated page loads, session would be better for truck no.
    # Let's stick to passing it back for the re-render after POST for now.
    return render_template("index.html",
                           result_message=result_val, 
                           error_message=error_val,
                           truck_numbers_master=truck_master_list,
                           submitted_truck_select=selected_truck, # Pass back what was selected
                           submitted_truck_other=other_truck_number if selected_truck == "_other_" else "" # Pass back other only if "Other" was chosen
                           )

@app.route("/trucks", methods=["GET"])
def manage_trucks():
    truck_master_list = get_truck_master_list_from_db()
    return render_template("trucks.html", trucks_master=truck_master_list)

@app.route("/trucks/add", methods=["POST"])
def add_truck_to_master():
    new_truck_number = request.form.get("new_truck_number", "").strip().upper()
    if not new_truck_number: flash("Truck number cannot be empty.", "error")
    else:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        try:
            cur.execute("INSERT INTO truck_master_list (truck_number) VALUES (?)", (new_truck_number,))
            conn.commit()
            flash(f"Truck '{new_truck_number}' added to master list!", "success")
        except sqlite3.IntegrityError: flash(f"Truck number '{new_truck_number}' already exists in the master list.", "error")
        except Exception as e: flash(f"Error adding truck to master list: {e}", "error")
        finally: conn.close()
    return redirect(url_for('manage_trucks'))

@app.route("/trucks/delete/<int:truck_id>", methods=["POST"])
def delete_truck_from_master(truck_id):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        cur.execute("SELECT truck_number FROM truck_master_list WHERE id = ?", (truck_id,))
        truck = cur.fetchone()
        if truck:
            cur.execute("DELETE FROM truck_master_list WHERE id = ?", (truck_id,))
            conn.commit()
            flash(f"Truck '{truck[0]}' deleted from master list!", "success")
        else: flash("Truck not found in master list.", "error")
    except Exception as e: flash(f"Error deleting truck from master list: {e}", "error")
    finally: conn.close()
    return redirect(url_for('manage_trucks'))

@app.route("/history")
def history():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("SELECT serial, timestamp, truck_number, system_timestamp FROM scans ORDER BY system_timestamp DESC, timestamp DESC")
    rows = cur.fetchall()
    conn.close()
    return render_template("history.html", rows=rows)

@app.route("/delete", methods=["POST"])
def delete_scans():
    serials_to_delete = request.form.getlist("delete")
    if serials_to_delete:
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        cur.executemany("DELETE FROM scans WHERE serial = ?", [(s,) for s in serials_to_delete])
        conn.commit(); conn.close()
        flash(f"{len(serials_to_delete)} scan(s) deleted.", "success")
    else: flash("No scans selected for deletion.", "info")
    return redirect(url_for('history'))

@app.route("/delete_all_scans", methods=["POST"])
def delete_all_scans():
    conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
    cur.execute("DELETE FROM scans"); count = cur.rowcount
    conn.commit(); conn.close()
    flash(f"All {count} scan records have been deleted.", "warning")
    return redirect(url_for('history'))

@app.route("/export", methods=["GET"])
def export_options_page():
    return render_template("export.html")

@app.route("/download_report", methods=["POST"])
def download_excel_report():
    price_per_load_str = request.form.get("price_per_load", "0").strip()
    try:
        price_per_load = float(price_per_load_str)
        if price_per_load < 0: price_per_load = 0; flash("Price non-negative. Using 0.", "warning")
    except ValueError:
        price_per_load = 0; flash("Invalid price. Using 0 for calculations.", "warning")

    conn = sqlite3.connect(DB_FILE); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    cur.execute("SELECT serial, timestamp, system_timestamp, truck_number FROM scans ORDER BY truck_number, system_timestamp ASC, timestamp ASC")
    all_scans_raw = cur.fetchall(); conn.close()
    
    output = io.BytesIO(); workbook = openpyxl.Workbook(); workbook.remove(workbook.active)
    
    f11, fb11, fb12 = Font(name='Calibri',size=11), Font(name='Calibri',size=11,bold=True), Font(name='Calibri',size=12,bold=True)
    hf_style = Font(name='Calibri',bold=True,color="FFFFFF")
    h_fill, s_fill, gt_fill = PatternFill(start_color="4F81BD",fill_type="solid"), PatternFill(start_color="F2F2F2",fill_type="solid"), PatternFill(start_color="C0C0C0",fill_type="solid")
    ca, la, ra = Alignment(horizontal="center",vertical="center"), Alignment(horizontal="left",vertical="center"), Alignment(horizontal="right",vertical="center")
    cf_str,int_str = '$#,##0.00', '#,##0'
    thin_s_obj, thin_bdr_style = Side(style='thin',color="BFBFBF"), Border(left=Side(style='thin',color="BFBFBF"),right=Side(style='thin',color="BFBFBF"),top=Side(style='thin',color="BFBFBF"),bottom=Side(style='thin',color="BFBFBF"))
    tbb_style = Border(top=thin_s_obj,bottom=thin_s_obj)

    # Sheet 1: All Scans Detailed
    s1 = workbook.create_sheet(title="All Scans Detailed")
    hdrs1=["Serial No","QR Timestamp","System Timestamp","Truck Number","Load Value"]
    for cn,ht in enumerate(hdrs1,1): c=s1.cell(row=1,column=cn,value=ht);c.font=hf_style;c.fill=h_fill;c.alignment=ca;c.border=thin_bdr_style
    
    sbt=defaultdict(list)
    for sr in all_scans_raw:sbt[sr["truck_number"]if sr["truck_number"]and sr["truck_number"].strip()else "Unassigned"].append(dict(sr))
    s_tks=sorted(sbt.keys(),key=lambda x:(x=="Unassigned",x.lower()if x!="Unassigned"else ""))
    
    er,gtrd,gtvd=2,0,0.0
    for tk in s_tks:
        tsl,nsl,tsv = sbt[tk],len(sbt[tk]),len(sbt[tk])*price_per_load
        gtrd+=nsl;gtvd+=tsv
        for sd in tsl:
            tnd=sd["truck_number"]if sd["truck_number"]else"N/A"
            c1=s1.cell(row=er,column=1,value=sd["serial"]);c1.font=f11;c1.alignment=la;c1.border=thin_bdr_style
            c2=s1.cell(row=er,column=2,value=sd["timestamp"]);c2.font=f11;c2.alignment=ca;c2.border=thin_bdr_style
            c3=s1.cell(row=er,column=3,value=sd["system_timestamp"]);c3.font=f11;c3.alignment=ca;c3.border=thin_bdr_style
            c4=s1.cell(row=er,column=4,value=tnd);c4.font=f11;c4.alignment=ca;c4.border=thin_bdr_style
            c5=s1.cell(row=er,column=5,value=price_per_load);c5.font=f11;c5.number_format=cf_str;c5.alignment=ra;c5.border=thin_bdr_style
            er+=1
        s1.merge_cells(start_row=er,start_column=1,end_row=er,end_column=3) # Merge A, B, C
        slc=s1.cell(row=er,column=1,value=f"Subtotal for Truck {tk}");slc.font=fb11;slc.alignment=ra;slc.fill=s_fill
        for ci in range(1,4):mc=s1.cell(row=er,column=ci);mc.fill=s_fill;mc.border=tbb_style
        src=s1.cell(row=er,column=4,value=nsl);src.font=fb11;src.alignment=ca;src.number_format=int_str;src.fill=s_fill;src.border=tbb_style
        svc=s1.cell(row=er,column=5,value=tsv);svc.font=fb11;svc.alignment=ra;svc.number_format=cf_str;svc.fill=s_fill;svc.border=tbb_style;er+=1
    if all_scans_raw:
        er+=1;s1.merge_cells(start_row=er,start_column=1,end_row=er,end_column=3) # Merge A,B,C
        gtlc=s1.cell(row=er,column=1,value="GRAND TOTALS");gtlc.font=fb12;gtlc.alignment=ra;gtlc.fill=gt_fill
        for ci in range(1,4):mc=s1.cell(row=er,column=ci);mc.fill=gt_fill;mc.border=tbb_style
        gtrc=s1.cell(row=er,column=4,value=gtrd);gtrc.font=fb12;gtrc.alignment=ca;gtrc.number_format=int_str;gtrc.fill=gt_fill;gtrc.border=tbb_style
        gtvc=s1.cell(row=er,column=5,value=gtvd);gtvc.font=fb12;gtvc.alignment=ra;gtvc.number_format=cf_str;gtvc.fill=gt_fill;gtvc.border=tbb_style
    
    for i, cl in enumerate(get_column_letter(idx + 1) for idx in range(s1.max_column)):
        ml = 0
        for ri in range(1, s1.max_row + 1):
            cell = s1.cell(row=ri, column=i + 1)
            cv = None; is_merged = any(mr.min_col <= cell.column <= mr.max_col and mr.min_row <= cell.row <= mr.max_row for mr in s1.merged_cells.ranges)
            if is_merged:
                for mro_loop in s1.merged_cells.ranges: 
                    if cell.coordinate in mro_loop: cv = s1.cell(row=mro_loop.min_row, column=mro_loop.min_col).value; break
            else: cv = cell.value
            if cv is not None: ml = max(ml, len(str(cv)))
        header_name_for_padding = hdrs1[i] if i < len(hdrs1) else "" 
        pad = 4 if header_name_for_padding in ["Load Value", "Total Receipts", "Truck Number"] else 2
        if header_name_for_padding == "QR Timestamp" or header_name_for_padding == "System Timestamp": pad = 5
        s1.column_dimensions[cl].width = (ml + pad if ml > 0 else 12)

    # Sheet 2: Summary by Truck
    s2 = workbook.create_sheet(title="Summary by Truck");hdrs2=["Truck Number","Total Receipts","Total Value"]
    for cn,ht in enumerate(hdrs2,1): c=s2.cell(row=1,column=cn,value=ht);c.font=hf_style;c.fill=h_fill;c.alignment=ca;c.border=thin_bdr_style
    tsd2 = defaultdict(lambda: {"count": 0})
    for s_scan_item in all_scans_raw:
        truck_key_for_summary = s_scan_item["truck_number"] if s_scan_item["truck_number"] and s_scan_item["truck_number"].strip() else "Unassigned"
        tsd2[truck_key_for_summary]["count"] += 1
    s_tks2=sorted(tsd2.keys(),key=lambda x:(x=="Unassigned",x.lower()if x!="Unassigned"else""));er2,trss,tvss=2,0,0.0
    for tk_s2 in s_tks2:
        cnt,tvft=tsd2[tk_s2]["count"],tsd2[tk_s2]["count"]*price_per_load;trss+=cnt;tvss+=tvft
        c1s=s2.cell(row=er2,column=1,value=tk_s2);c1s.font=f11;c1s.alignment=la;c1s.border=thin_bdr_style
        c2s=s2.cell(row=er2,column=2,value=cnt);c2s.font=f11;c2s.number_format=int_str;c2s.alignment=ca;c2s.border=thin_bdr_style
        c3s=s2.cell(row=er2,column=3,value=tvft);c3s.font=f11;c3s.number_format=cf_str;c3s.alignment=ra;c3s.border=thin_bdr_style;er2+=1
    if tsd2:
        er2+=1;gtls=s2.cell(row=er2,column=1,value="GRAND TOTAL");gtls.font=fb12;gtls.alignment=ra;gtls.fill=gt_fill;gtls.border=tbb_style
        gtrs=s2.cell(row=er2,column=2,value=trss);gtrs.font=fb12;gtrs.number_format=int_str;gtrs.alignment=ca;gtrs.fill=gt_fill;gtrs.border=tbb_style
        gtvs=s2.cell(row=er2,column=3,value=tvss);gtvs.font=fb12;gtvs.number_format=cf_str;gtvs.alignment=ra;gtvs.fill=gt_fill;gtvs.border=tbb_style
    for i, cl in enumerate(get_column_letter(idx + 1) for idx in range(s2.max_column)):
        ml = 0
        for ri in range(1, s2.max_row + 1): cv = s2.cell(row=ri, column=i + 1).value
        if cv is not None: ml = max(ml, len(str(cv)))
        pad = 4 if hdrs2[i] in ["Total Value", "Total Receipts"] else 2
        s2.column_dimensions[cl].width = (ml + pad if ml > 0 else 15)

    # Sheet 3: Overall Summary
    s3 = workbook.create_sheet(title="Overall Summary");hdrs3=["Description","Value"]
    for cn,ht in enumerate(hdrs3,1):c=s3.cell(row=1,column=cn,value=ht);c.font=hf_style;c.fill=h_fill;c.alignment=ca if cn==2 else la;c.border=thin_bdr_style
    trov,tvov=len(all_scans_raw),len(all_scans_raw)*price_per_load
    sdr=[["Total Receipts Scanned",trov],[f"Price Per Load Used",price_per_load],["Total Calculated Value",tvov]];er3=2
    for i_r, dr_item in enumerate(sdr):
        islr,cfs=(i_r==len(sdr)-1),fb12 if(i_r==len(sdr)-1)else f11
        dc=s3.cell(row=er3,column=1,value=dr_item[0]);dc.font=cfs;dc.alignment=la;dc.border=thin_bdr_style
        vc=s3.cell(row=er3,column=2,value=dr_item[1]);vc.font=cfs;vc.alignment=ra;vc.border=thin_bdr_style
        if "Price"in dr_item[0]or "Value"in dr_item[0]:vc.number_format=cf_str
        else:vc.number_format=int_str
        er3+=1
    s3.column_dimensions['A'].width=35;s3.column_dimensions['B'].width=20

    workbook.save(output); output.seek(0)
    return send_file(output,as_attachment=True,download_name="receipt_scan_report_styled.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    flask_port = 5001
    print(f"---------------------------------------------------------------------")
    print(f"Flask app starting on http://127.0.0.1:{flask_port} (or http://0.0.0.0:{flask_port})")
    print("To enable camera on mobile Chrome when accessing via this machine's IP address,")
    print("you generally need to serve the application over HTTPS.")
    print("For development, consider using a tunneling service like ngrok:")
    print(f"  1. Run this Flask app (python app.py)")
    print(f"  2. In a SEPARATE terminal, run: ngrok http {flask_port}")
    print(f"  3. Access the https://<random-string>.ngrok.io URL provided by ngrok on your mobile.")
    print(f"Database file: {DB_FILE}")
    print(f"---------------------------------------------------------------------")
    
    app.run(host="0.0.0.0", port=flask_port, debug=True, use_reloader=True)